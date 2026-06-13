"""
Phase 2: CSV and Excel upload endpoint.
PDF bank statement parsing is deferred to Phase 5 / future work
(bank PDF formats vary widely and require per-bank templates).
"""
import io
import csv
from datetime import date as date_type
from fastapi import APIRouter, Depends, File, UploadFile, HTTPException
from sqlalchemy.orm import Session
from ..core.database import get_db
from ..core.deps import get_current_user
from ..core.config import settings
from ..models.user import User
from ..models.transaction import Transaction
from ..services.categorizer import categorize, extract_merchant, ai_categorize

router = APIRouter(prefix="/upload", tags=["upload"])

ALLOWED_TYPES = {
    "text/csv", "application/vnd.ms-excel",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
EXPECTED_COLS = {"date", "description", "amount", "type"}

def _parse_date(val: str) -> date_type:
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y"):
        try:
            from datetime import datetime
            return datetime.strptime(val.strip(), fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognised date format: {val}")

def _rows_from_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [row for row in reader]

def _rows_from_excel(content: bytes) -> list[dict]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(c.value or "").strip().lower().replace(" ", "_") for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
    return rows

async def _process_rows(rows: list[dict], user_id: int, db: Session) -> dict:
    loaded = 0
    skipped = 0
    for row in rows:
        # normalise column names to lowercase + underscores
        row = {k.strip().lower().replace(" ", "_"): v for k, v in row.items()}
        if not all(col in row for col in EXPECTED_COLS):
            skipped += 1
            continue
        try:
            txn_date = _parse_date(str(row["date"]))
        except ValueError:
            skipped += 1
            continue
        try:
            amount = float(str(row["amount"]).replace(",", "").replace("₹", "").strip())
        except ValueError:
            skipped += 1
            continue

        desc     = str(row.get("description", "")).strip()
        txn_type = str(row.get("type", "debit")).strip().lower()
        source   = str(row.get("source", "")).strip() or None
        account  = str(row.get("account", "")).strip() or None
        merchant = str(row.get("merchant", "")).strip() or extract_merchant(desc)
        category = str(row.get("category", "")).strip()

        if not category:
            category = categorize(desc, merchant)
            if category == "Miscellaneous" and settings.OPENAI_API_KEY:
                category = await ai_categorize(desc, merchant,
                                               settings.OPENAI_API_KEY,
                                               settings.OPENAI_BASE_URL,
                                               settings.OPENAI_MODEL)

        db.add(Transaction(
            user_id=user_id, date=txn_date,
            description=desc, amount=amount, type=txn_type,
            source=source, account=account,
            category=category, merchant=merchant or None,
        ))
        loaded += 1

    db.commit()
    return {"loaded": loaded, "skipped": skipped}

@router.post("", summary="Upload CSV or Excel transaction file")
async def upload_file(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    content_type = file.content_type or ""
    filename     = (file.filename or "").lower()

    if not (filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".xls")):
        raise HTTPException(status_code=400, detail="Only .csv and .xlsx/.xls files are supported. PDF parsing is a future feature.")

    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:   # 10 MB guard
        raise HTTPException(status_code=413, detail="File too large (max 10 MB)")

    if filename.endswith(".csv"):
        rows = _rows_from_csv(raw)
    else:
        rows = _rows_from_excel(raw)

    if not rows:
        raise HTTPException(status_code=400, detail="File appears empty")

    result = await _process_rows(rows, current_user.id, db)
    return {
        "message": f"Upload complete. {result['loaded']} transactions imported, {result['skipped']} rows skipped.",
        **result,
    }
